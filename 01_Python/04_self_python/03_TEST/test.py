from datetime import datetime as dt
from openpyxl import load_workbook
from openpyxl.styles import Alignment, NamedStyle, Font, PatternFill
from openpyxl.styles.colors import Color
import os
import pandas as pd
import stat


root = os.path.dirname(__file__) 
file = os.path.join(root, "data/02/kokyaku_daicho.xlsx")
df = pd.read_excel(    
    # openpyex1が必要（issing optional dependency 'openpyxl'.  Use pip or conda to install openpyxl.）
    file, sheet_name=0, header=0
    )

print(df.head())
print(len(df))
print(df.columns)
print(df.index)


def str_to_datetime(str):
    '''文字列をdatetime型にする。
    引数:yyyy/mm/ddで受け取る'''
    tdatetime = dt.strptime(str, '%Y/%m/%d')
    return tdatetime

def strToday():
    '''現在日付を文字列で取得する。戻り値:'''
    now = dt.now()
    strToday = now.strptime(str, '%Y%m%d')
    return strToday


def chmod_W_OK(f):
    '''ファイルの読み取り専用を外す'''
    if os.path.exists(f):
        os.chmod(path=f, mode=stat.S_IWRITE)

def chmod_W_NG(f):
    '''ファイル読み取り専用にする'''
    if os.path.exists(f):
        os.chmod(path=f, mode=stat.S_IREAD)      


def output_xslx(df, xlsxFile, header):
    '''エクセルファイル出力。書式設定込み
    header例:A1:F1
    '''
    try:
        df.to_excel(
            xlsxFile, sheet_name=0, 
            header=True, index=True, startrow=0, startcol=0, na_rap="null"
        )

        rows = len(df)
        wb = load_workbook(xlsxFile)
        ws = wb.worksheets[0]

        header_style = NamedStyle(
            name='header_style',
            font=Font(name='Arial', size=12, bold=True),
            fill=PatternFill(patternType='solid', fgColor=Color(rgb='bce2e8')),    # 塗りつぶしの設定
            alignment=Alignment(wrap_text=True, shrink_to_fit=True, horizontal='general', vertical='center')
        )
        for row in ws[header]:             # ヘッダーを仮にA1:F1とする
            for cell in row:
                cell.style=header_style
        ws.column_dimensions['A'].width=20

        return True
    
    except:
        return False